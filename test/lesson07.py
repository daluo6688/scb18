"""
接口自动化测试
1、excel测试用例准备OK，代码可以自动读取用例数据-------read_case(filename,sheetname)
2、执行接口测试，得到测试结果   api_fun(url,data)
3、断言，响应结果==测试结果   --通过/不通过
4、写到最终执行通过与否的结果---excel表格 ------write_result(filename,sheetname,row,column,final_result)

"""
import requests
import openpyxl

#读取测试用例
def read_case(filename,sheetname):
    wb = openpyxl.load_workbook(filename) #加载工作簿，打开一个excel文档
    sheet = wb[sheetname]  #打开某一个表单
    row_max = sheet.max_row #获取最大行数
    case_list = []  #新建空列表，存放for循环依次读取到的测试用例数据
    for i in range(2,row_max+1):
        data_dict = dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value, #获取url
        data = sheet.cell(row=i,column=6).value ,#获取data值
        expect = sheet.cell(row=i,column=7).value,  #获取期望
        )
        case_list.append(data_dict) #把每一行读取到的测试用例数据生成的字典，追加到list中
    return case_list

#执行测试用例
def api_fun(url,data):
    headers = {'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'} #请求头
    res = requests.post(url=url,json=data,headers=headers).json()  #调用post，用变量接收返回值
    return res

#写入测试结果
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename) #加载工作簿，打开一个excel文档
    sheet = wb[sheetname]  #打开某一个表单
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)

#断言：实际结果==预期结果
#封装成一个函数
def execute_fun(filename,sheetname):
    cases = read_case(filename,sheetname)  #调用函数，设置变量接收返回值
    for case in cases:
        case_id = case['case_id']  #获取用例编号
        url_case = case['url'] #获取URL的值
        data = eval(case['data'])  #字符串==字典
        expect = eval(case['expect']) #预期结果
        # print(case_id,expect)
        expect_msg = expect['msg']  #预期结果中的msg
        # print(expect_msg)
        real_result = api_fun(url=url_case,data=data)
        real_msg = real_result['msg']  #获取实际结果中的msg
        print('期望结果为：{}'.format(expect_msg))
        print('实际结果为：{}'.format(real_msg))
        # print(case_id,expect_msg,real_msg)
        if expect_msg==real_msg:
            print('{}用例执行通过！'.format(case_id))
            final_re = 'passed'
        else:
            print('{}用例执行不通过！'.format(case_id))
            final_re = 'Failed'
        write_result(filename,sheetname,case_id+1,8,final_re)
        print('*'*25)
#调用这个函数
execute_fun('../test_data/test_case_api.xlsx', 'register')
execute_fun('../test_data/test_case_api.xlsx', 'login')
#eval()---运行被字符串包裹的表达式
#"mobile_phone": "13540342011","pwd":"lemon666"
# print(eval('{'mobile_phone':'13540342011','pwd':'12345678','type':1,'reg_name':'34254sdfs'}'))


